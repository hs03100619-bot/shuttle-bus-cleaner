import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
import os

st.title("셔틀버스 수요조사 자동 정리 프로그램")

st.write("수요조사 결과 파일과 빈 양식 엑셀을 업로드하세요.")

# 파일 업로드 섹션
survey_file = st.file_uploader("1. 수요조사 파일 업로드 (엑셀 또는 CSV)", type=['xlsx', 'csv'])
template_file = st.file_uploader("2. 셔틀 양식 엑셀 파일 업로드", type=['xlsx'])

if survey_file and template_file:
    if st.button("데이터 정리 시작"):
        try:
            # 1. 수요조사 데이터 읽기
            if survey_file.name.endswith('.csv'):
                survey_df = pd.read_csv(survey_file)
            else:
                survey_df = pd.read_excel(survey_file)
            
            id_col = [col for col in survey_df.columns if '교번' in col][0]
            dep_col = [col for col in survey_df.columns if '출교' in col][0]
            ret_col = [col for col in survey_df.columns if '귀교' in col][0]
            
            survey_dict = {}
            for idx, row in survey_df.iterrows():
                student_id = str(row[id_col]).strip()
                dep = str(row[dep_col]).strip()
                ret = str(row[ret_col]).strip()
                survey_dict[student_id] = (dep, ret)

            def map_val(v):
                v = str(v).strip().lower()
                if v == 'o': return '○'
                if v == 'x': return 'X'
                return v

            # 2. 양식 파일에 데이터 쓰기
            wb = openpyxl.load_workbook(template_file)
            ws = wb.active

            for row_idx in range(3, ws.max_row + 1):
                student_id_cell = ws.cell(row=row_idx, column=1)
                student_id = str(student_id_cell.value).strip()
                
                if student_id == 'None' or student_id == '':
                    continue
                    
                if student_id in survey_dict:
                    dep, ret = survey_dict[student_id]
                    ws.cell(row=row_idx, column=5).value = map_val(dep)
                    ws.cell(row=row_idx, column=6).value = map_val(ret)
                else:
                    ws.cell(row=row_idx, column=5).value = '답변X'
                    ws.cell(row=row_idx, column=6).value = '답변X'

            # 3. 결과물 저장
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            st.success("데이터 정리가 완료되었습니다. 아래 버튼을 눌러 다운로드하세요.")
            st.download_button(
                label="정리본 엑셀 다운로드",
                data=output,
                file_name="셔틀_정리본.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"오류가 발생했습니다. 파일 양식을 확인해주세요: {e}")

# --- 하단 사용 방법 안내 섹션 추가 ---
st.markdown("---")
st.header("💡 프로그램 사용 방법")

st.subheader("1단계: 파일 업로드하기")
st.write("네이버폼, 구글폼 등에서 내려받은 **수요조사 결과 파일(Excel 또는 CSV)**을 첫 번째 칸에 업로드하고, 작성할 빈 **셔틀 양식 파일**을 두 번째 칸에 업로드합니다.")

img1_path = "화면 캡처 2026-07-11 100349.png"
if os.path.exists(img1_path):
    st.image(img1_path, caption="[참고] 파일 업로드 및 준비 화면", use_container_width=True)
else:
    st.caption(f"({img1_path} 파일이 폴더 내에 존재하지 않아 이미지를 표시할 수 없습니다.)")

img2_path = "화면 캡처 2026-07-11 100318.png"
if os.path.exists(img2_path):
    st.image(img2_path, caption="[참고] 데이터 정리 완료 및 다운로드 화면", use_container_width=True)
else:
    st.caption(f"({img2_path} 파일이 폴더 내에 존재하지 않아 이미지를 표시할 수 없습니다.)")

st.subheader("2단계: 데이터 정리 및 다운로드")
st.write("**[데이터 정리 시작]** 버튼을 누르면 프로그램이 교번을 기준으로 명단을 매칭하여 탑승 여부('○', 'X')를 입력합니다. 설문에 참여하지 않은 인원은 자동으로 **'답변X'**로 처리됩니다. 작업이 완료되면 **[정리본 엑셀 다운로드]** 버튼을 눌러 결과 파일을 저장합니다.")


